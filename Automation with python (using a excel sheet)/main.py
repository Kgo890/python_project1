import openpyxl

inv_file = openpyxl.load_workbook("Inventory.xlsx")
product_list = inv_file["Sheet1"]

products_per_supplier = {}
total_value_per_supplier = {}
product_under_10 = {}
# use max_row to read all the row with in the Excel sheet
# (2,...) in the range() helps bypass the titles in first row
# + 1 makes sure that the last row is read

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory_product = int(product_list.cell(product_row, 2).value)
    price = float(product_list.cell(product_row, 3).value)
    product_num = int(product_list.cell(product_row, 1).value)
    inventory_price = product_list.cell(product_row, 5)

    # calculating the number of products per supplier
    if supplier_name in products_per_supplier:
        current_num_products = products_per_supplier.get(supplier_name)
        products_per_supplier[supplier_name] = current_num_products + 1

    else:
        products_per_supplier[supplier_name] = 1

    # calculation total value of the inventory
    if supplier_name in total_value_per_supplier:
        current_total = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = round(current_total + inventory_product * price, 2)
    else:
        total_value_per_supplier[supplier_name] = round(inventory_product * price, 2)

    # logic products with inventory less than ten
    if inventory_product < 10:

        product_under_10[product_num] = inventory_product

    # add value for total inventory price
    inventory_price.value = inventory_product * price

    # save file
    inv_file.save("new_inventory.xlsx")

print(products_per_supplier)
print(total_value_per_supplier)
print(product_under_10)
